from typing import Dict, Any, Tuple
import os
from datetime import datetime
from flask import current_app
from app.repositories.student_repo import StudentRepository
from app.models.student import Student
from .base_service import BaseService


class ReportService(BaseService):
    def __init__(self):
        self.student_repo = StudentRepository()
        super().__init__(self.student_repo)
    
    def generate_student_report(self, format: str = 'pdf', search_term: str = None, **filters) -> Tuple[Dict[str, Any], int]:
        """Generate student report in specified format"""
        try:
            # Get student data
            students = self.student_repo.get_students_with_details()
            
            # Apply filters
            if search_term:
                search = search_term.lower()
                students = [
                    s for s in students 
                    if (search in s['student_id'].lower() or
                        search in s['user'].get('first_name', '').lower() or
                        search in s['user'].get('last_name', '').lower() or
                        search in s['user'].get('email', '').lower())
                ]
            
            for key, value in filters.items():
                if value is not None:
                    students = [s for s in students if str(s.get(key, '')).lower() == str(value).lower()]
            
            if format == 'pdf':
                file_path = self._generate_pdf_report(students)
            else:
                file_path = self._generate_excel_report(students)
            
            filename = f"student_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.{format if format != 'excel' else 'xlsx'}"
            
            return {
                'status': 'success',
                'file_path': file_path,
                'filename': filename
            }, 200
            
        except Exception as e:
            current_app.logger.error(f"Error generating report: {str(e)}")
            return {
                'status': 'error',
                'message': f'Failed to generate report: {str(e)}'
            }, 500
    
    def generate_student_profile(self, student_id: int, format: str = 'pdf') -> Tuple[Dict[str, Any], int]:
        """Generate individual student profile report"""
        try:
            student = self.student_repo.get_student_with_details(student_id)
            
            if not student:
                return {
                    'status': 'error',
                    'message': 'Student not found'
                }, 404
            
            file_path = self._generate_student_profile_pdf(student)
            filename = f"student_profile_{student['student_id']}_{datetime.now().strftime('%Y%m%d')}.pdf"
            
            return {
                'status': 'success',
                'file_path': file_path,
                'filename': filename
            }, 200
            
        except Exception as e:
            current_app.logger.error(f"Error generating student profile: {str(e)}")
            return {
                'status': 'error',
                'message': f'Failed to generate profile: {str(e)}'
            }, 500
    
    def get_student_analytics(self) -> Tuple[Dict[str, Any], int]:
        """Get student statistics and analytics"""
        try:
            from app.models.role import Role
            from app import db
            
            total_students = self.student_repo.model.query.count()
            
            # Gender distribution
            gender_stats = db.session.query(
                Student.gender,
                db.func.count(Student.id)
            ).group_by(Student.gender).all()
            
            # Recent admissions (last 30 days)
            from datetime import timedelta
            thirty_days_ago = datetime.now().date() - timedelta(days=30)
            recent_admissions = Student.query.filter(
                Student.admission_date >= thirty_days_ago
            ).count()
            
            return {
                'status': 'success',
                'analytics': {
                    'total_students': total_students,
                    'gender_distribution': {gender: count for gender, count in gender_stats if gender},
                    'recent_admissions': recent_admissions
                }
            }, 200
            
        except Exception as e:
            current_app.logger.error(f"Error fetching analytics: {str(e)}")
            return {
                'status': 'error',
                'message': 'Failed to fetch analytics'
            }, 500
    
    def _generate_pdf_report(self, students: list) -> str:
        """Generate PDF report for students list"""
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            
            # Create reports directory if it doesn't exist
            reports_dir = current_app.config.get('REPORTS_FOLDER')
            pdf_dir = os.path.join(reports_dir, 'pdf')
            os.makedirs(pdf_dir, exist_ok=True)
            
            filename = f"students_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            filepath = os.path.join(pdf_dir, filename)
            
            doc = SimpleDocTemplate(filepath, pagesize=A4)
            elements = []
            
            # Styles
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                textColor=colors.HexColor('#1a1a1a'),
                spaceAfter=30,
                alignment=1  # Center
            )
            
            # Add title
            title = Paragraph("Student Management System - Student Report", title_style)
            elements.append(title)
            elements.append(Spacer(1, 0.3*inch))
            
            # Add generation date
            date_text = Paragraph(
                f"Generated on: {datetime.now().strftime('%B %d, %Y at %H:%M')}",
                styles['Normal']
            )
            elements.append(date_text)
            elements.append(Spacer(1, 0.3*inch))
            
            # Prepare table data
            data = [['Student ID', 'Name', 'Email', 'Gender', 'Phone', 'Admission Date']]
            
            for student in students:
                user = student.get('user', {})
                full_name = f"{user.get('first_name', '')} {user.get('last_name', '')}".strip()
                data.append([
                    student.get('student_id', 'N/A'),
                    full_name or 'N/A',
                    user.get('email', 'N/A'),
                    student.get('gender', 'N/A'),
                    student.get('phone', 'N/A'),
                    student.get('admission_date', 'N/A')
                ])
            
            # Create table
            table = Table(data, repeatRows=1)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4a90e2')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
            ]))
            
            elements.append(table)
            
            # Build PDF
            doc.build(elements)
            
            return filepath
            
        except ImportError:
            # If reportlab is not installed, create a simple text file
            current_app.logger.warning("reportlab not installed, generating text report instead")
            return self._generate_text_report(students, 'pdf')
    
    def _generate_excel_report(self, students: list) -> str:
        """Generate Excel report for students list"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            
            # Create reports directory
            reports_dir = current_app.config.get('REPORTS_FOLDER')
            excel_dir = os.path.join(reports_dir, 'excel')
            os.makedirs(excel_dir, exist_ok=True)
            
            filename = f"students_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = os.path.join(excel_dir, filename)
            
            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Students Report"
            
            # Headers
            headers = ['Student ID', 'First Name', 'Last Name', 'Email', 'Gender', 'Phone', 'Date of Birth', 'Admission Date']
            ws.append(headers)
            
            # Style headers
            header_fill = PatternFill(start_color="4A90E2", end_color="4A90E2", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            
            # Add data
            for student in students:
                user = student.get('user', {})
                ws.append([
                    student.get('student_id', ''),
                    user.get('first_name', ''),
                    user.get('last_name', ''),
                    user.get('email', ''),
                    student.get('gender', ''),
                    student.get('phone', ''),
                    student.get('date_of_birth', ''),
                    student.get('admission_date', '')
                ])
            
            # Adjust column widths
            for column in ws.columns:
                max_length = 0
                column = list(column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column[0].column_letter].width = adjusted_width
            
            wb.save(filepath)
            return filepath
            
        except ImportError:
            current_app.logger.warning("openpyxl not installed, generating text report instead")
            return self._generate_text_report(students, 'excel')
    
    def _generate_student_profile_pdf(self, student: dict) -> str:
        """Generate PDF profile for individual student"""
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            
            reports_dir = current_app.config.get('REPORTS_FOLDER')
            pdf_dir = os.path.join(reports_dir, 'pdf')
            os.makedirs(pdf_dir, exist_ok=True)
            
            filename = f"student_profile_{student['student_id']}_{datetime.now().strftime('%Y%m%d')}.pdf"
            filepath = os.path.join(pdf_dir, filename)
            
            doc = SimpleDocTemplate(filepath, pagesize=A4)
            elements = []
            styles = getSampleStyleSheet()
            
            # Title
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                textColor=colors.HexColor('#1a1a1a'),
                spaceAfter=20,
                alignment=1
            )
            title = Paragraph("Student Profile", title_style)
            elements.append(title)
            elements.append(Spacer(1, 0.3*inch))
            
            # Student information
            user = student.get('user', {})
            info_data = [
                ['Student ID:', student.get('student_id', 'N/A')],
                ['Name:', f"{user.get('first_name', '')} {user.get('last_name', '')}"],
                ['Email:', user.get('email', 'N/A')],
                ['Date of Birth:', student.get('date_of_birth', 'N/A')],
                ['Gender:', student.get('gender', 'N/A')],
                ['Phone:', student.get('phone', 'N/A')],
                ['Address:', student.get('address', 'N/A')],
                ['Admission Date:', student.get('admission_date', 'N/A')],
                ['Status:', 'Active' if user.get('is_active', False) else 'Inactive']
            ]
            
            info_table = Table(info_data, colWidths=[2*inch, 4*inch])
            info_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
                ('ALIGN', (1, 0), (1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 11),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            elements.append(info_table)
            
            doc.build(elements)
            return filepath
            
        except ImportError:
            current_app.logger.warning("reportlab not installed")
            return self._generate_text_report([student], 'profile')
    
    def _generate_text_report(self, students: list, report_type: str) -> str:
        """Fallback text report generation"""
        reports_dir = current_app.config.get('REPORTS_FOLDER')
        os.makedirs(reports_dir, exist_ok=True)
        
        filename = f"students_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
        filepath = os.path.join(reports_dir, filename)
        
        with open(filepath, 'w') as f:
            f.write("Student Management System - Report\n")
            f.write("=" * 80 + "\n\n")
            f.write(f"Generated on: {datetime.now().strftime('%B %d, %Y at %H:%M')}\n\n")
            
            for student in students:
                user = student.get('user', {})
                f.write(f"Student ID: {student.get('student_id', 'N/A')}\n")
                f.write(f"Name: {user.get('first_name', '')} {user.get('last_name', '')}\n")
                f.write(f"Email: {user.get('email', 'N/A')}\n")
                f.write(f"Gender: {student.get('gender', 'N/A')}\n")
                f.write(f"Phone: {student.get('phone', 'N/A')}\n")
                f.write(f"Admission Date: {student.get('admission_date', 'N/A')}\n")
                f.write("-" * 80 + "\n\n")
        
        return filepath